from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import pandas as pd 
import openpyxl 
import os
from inventory_analysis import InventoryRates
import tooltip
import configparser
from os import path


import logging
logging.basicConfig(filename='inventory_analysis.log',level=logging.DEBUG,format='%(asctime)s:%(levelname)s:%(message)s')
'''
window = Tk()

label = Label(window,text='Inventory Analysis')
#text = Text(window, cnf={'bg':'blue'})


def on_clear():
	print("Clear")

def on_submit():
	print("Submitting")

label = Label(window,text='Inventory Analysis')
#text = Text(window, cnf={'bg':'blue'})
button_clear = Button(window, text='OK', bg='blue', command=on_clear)
button_submit = Button(window, text='Submit', fg='green',command=on_submit)

button_clear.grid(row=1,column=1,padx=(2,10),pady=5)
button_submit.grid(row=1,column=2)

window.mainloop()
'''

window = Tk()
window.title("Inventory Analysis")
window.geometry("500x700")



class IAGui:
	def __init__(self, window, is_config):
		#window = Tk()
		#window.title("Inventory Analysis")
		#window.geometry=("500x500")
		

		#outputs
		#Label(window, text="Result").grid(row=5, column=1, sticky=W)
		#Label(window, text="Result 2").grid(row=6, column=1, sticky=W)
		if is_config:
			basepath = path.dirname(__file__)

			config_obj = configparser.ConfigParser()
			config_obj.read(path.join(basepath,'configfile.ini'))

			col_names = config_obj['column headers']
			
			self.data_type = StringVar(value=col_names['data_type'])
			self.product = StringVar(value=col_names['product'])
			self.product1 = StringVar(value=col_names['product1'])
			self.product2  = StringVar(value=col_names['product2'])
			self.facility = StringVar(value=col_names['facility'])
			self.facility1 = StringVar(value=col_names['facility1'])
			self.facility2 = StringVar(value=col_names['facility2'])
			self.date_field = StringVar(value=col_names['date_field'])
			self.issue_col = StringVar(value=col_names['issue_col'])
			self.soh_col = StringVar(value=col_names['soh_col'])

			other = config_obj['other']

			self.rates_window = IntVar(value=int(other['rates_window']))
			self.issue_0 = IntVar(value=int(other['issue_0']))
			self.soh_0 = IntVar(value=int(other['soh_0']))

			self.filename = ""
			self.column_names = ["N/A"]
			self.sheetname = StringVar()

		else:

			self.data_type = StringVar()
			self.product = StringVar()
			self.product1 = StringVar()
			self.product2  = StringVar()
			self.facility = StringVar()
			self.facility1 = StringVar()
			self.facility2 = StringVar()
			self.date_field = StringVar()
			self.rates_window = IntVar(value=12)
			self.issue_col = StringVar()
			self.issue_0 = IntVar(value=0)
			self.soh_col = StringVar()
			self.soh_0 = IntVar(value=0)
			self.filename = ""
			self.column_names = ["N/A"]
			self.sheetname = StringVar()


		self.label_file_explorer = Label(window,text = "Select an inventory data file to begin",fg = "blue")
		self.label_file_explorer.grid(row=1,column=2)

		file_button = Button(window,
						text = "Browse Files",
						command = self.browseFiles)
		file_button.grid(row=2,column=2)


		#window.mainloop()

	def browseFiles(self):
		self.filename = filedialog.askopenfilename(initialdir = "/",title = "Select a File", filetypes = (("Excel","*.xlsx"),("CSV","*.csv")))
		self.label_file_explorer.configure(text="File Opened: "+self.filename.split("/")[-1])
		logging.info('File selected: {}'.format(self.filename))

		if self.filename:

			if str(self.filename).endswith('.xlsx'):
				#wb = openpyxl.load_workbook(self.filename)
				#print(wb.sheetnames)
				wb = pd.ExcelFile(self.filename)
				print(wb.sheet_names)
				if len(wb.sheet_names) == 1:
					self.sheetname.set(wb.sheet_names[0])
				else:
					Label(window, text="Sheet").grid(row=1, column=1,sticky=E)
					ttk.Combobox(window, textvariable=self.sheetname, values=wb.sheet_names,justify=LEFT).grid(row=1,column=2, padx=(0,5))
				self.column_names = [x.strip() for x in list(pd.read_excel(self.filename,sheet_name = self.sheetname.get(),nrows=1).columns)]
			elif str(self.filename).endswith('.csv'):
				self.column_names = [x.strip() for x in list(pd.read_csv(self.filename,nrows=1).columns)]

		dataTypeLabel = Label(window, text="Data Type")
		dataTypeLabel.grid(row=3, column=1, sticky=NE)
		tooltip.Create(dataTypeLabel, "This is the data type")
		Label(window, text="Product").grid(row=4, column=1,sticky=E)
		Label(window, text="(Additional Product Info)").grid(row=5,column=1,sticky=E)
		Label(window, text="(Additional Product Info)").grid(row=6,column=1,sticky=NE)
		Label(window, text="Facility").grid(row=7, column=1, sticky=E)
		Label(window, text="(Additional Facility Info)").grid(row=8,column=1,sticky=E)
		Label(window, text="(Additional Facility Info)").grid(row=9,column=1,sticky=NE)
		Label(window, text="Date Field").grid(row=10, column=1, sticky=E)
		Label(window, text="Consumption Field").grid(row=11, column=1, sticky=E)
		Label(window, text="Stock on Hand Field").grid(row=12, column=1, sticky=NE)
		Label(window, text="Time Window for Rates").grid(row=13, column=1, sticky=E)

		#Label(window, text=None).grid(row=11, column=1, sticky=E)

		data_type_entry = ttk.Combobox(window, textvariable=self.data_type, values=['transactional','cumulative'],
			  justify=LEFT)
		data_type_entry.grid(row=3,column=2, padx=(0,5),pady=(0,50))
		print(self.column_names)
		ttk.Combobox(window, textvariable=self.product, values=self.column_names,
			  justify=LEFT).grid(row=4,column=2, padx=(0,5))
		ttk.Combobox(window, textvariable=self.product1, values=self.column_names,
			  justify=LEFT).grid(row=5,column=2, padx=(40,0))
		ttk.Combobox(window, textvariable=self.product2, values=self.column_names,
			  justify=LEFT).grid(row=6,column=2, padx=(40,0),pady=(0,50))
		ttk.Combobox(window, textvariable=self.facility, values=self.column_names,
			  justify=LEFT).grid(row=7,column=2, padx=(0,5))
		ttk.Combobox(window, textvariable=self.facility1, values=self.column_names,
			  justify=LEFT).grid(row=8,column=2, padx=(40,0))
		ttk.Combobox(window, textvariable=self.facility2, values=self.column_names,
			  justify=LEFT).grid(row=9,column=2, padx=(40,0),pady=(0,50))
		ttk.Combobox(window, textvariable=self.date_field, values=self.column_names,
			  justify=LEFT).grid(row=10,column=2, padx=(0,5))
		ttk.Combobox(window, textvariable=self.issue_col, values=self.column_names,
			  justify=LEFT).grid(row=11,column=2, padx=(0,5))
		Radiobutton(window,text="Blank fields = 0",variable=self.issue_0,value=1).grid(row=11,column=3,sticky=W)
		Radiobutton(window,text="Blank fields are missing",variable=self.issue_0,value=0).grid(row=11,column=4,sticky=W)
		ttk.Combobox(window, textvariable=self.soh_col, values=self.column_names,
			  justify=LEFT).grid(row=12,column=2, padx=(0,5),pady=(0,50))
		Radiobutton(window,text="Blank fields = 0",variable=self.soh_0,value=1).grid(row=12,column=3,sticky=W)
		Radiobutton(window,text="Blank fields are missing",variable=self.soh_0,value=0).grid(row=12,column=4,sticky=W)
		ttk.Combobox(window, textvariable=self.rates_window, values=list(range(6,25)),
			  justify=LEFT).grid(row=13,column=2, padx=(0,5))

		run_button = Button(window,
							text="Run Analysis",
							command = self.runIA)
		run_button.grid(row=15,column=2)
	
	def runIA(self):
		products = list(filter(None,[self.product.get(),self.product1.get(),self.product2.get()]))
		facilities = list(filter(None,[self.facility.get(),self.facility1.get(),self.facility2.get()]))
		i = InventoryRates(inv_type = self.data_type.get(), data_path = self.filename, date_col = self.date_field.get(), issue_col = self.issue_col.get(), soh_col = self.soh_col.get(), window= self.rates_window.get(),agg_cols=facilities+products,fac_cols = facilities, prod_cols=products)
		Label(window, text='Calculating Rates').grid(row=16,column=2)
		logging.debug('Product columns: {}'.format(products))
		logging.debug('Facility columns: {}'.format(facilities))
		logging.info('Inventory rates object instantiated')
		i.rolling_rates()
		logging.info('Inventory rates calculated')
		Label(window, text='Rates Calculated').grid(row=16,column=2)

		report_button = Button(window,
							text="Open Report in Excel",
							command = self.openReport(i))
		report_button.grid(row=17,column=2)
		num_prods = "Total Products: "+str(i.prod_table.shape[0])
		num_facs = "Total Facilities: "+str(i.fac_table.shape[0])
		Label(window, text=num_prods).grid(row=18,column=2)
		Label(window, text=num_facs).grid(row=19,column=2)
		#run_button.config(text="Re-run Analysis")
		config_button = Button(window,
							text="Save My Settings",
							command = self.createConfig)
		config_button.grid(row = 20,column=2)

	def openReport(self, i):
		
		#wb = openpyxl.load_workbook(filename = 'LMIS Analysis Report.xlsm', read_only = False, keep_vba = True)
		writer = pd.ExcelWriter('LMIS Analysis Report Data.xlsx')
		#writer.book = wb
		#writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
		i.rates_data.to_excel(writer, "Data",index=False)
		i.prod_table.to_excel(writer, "Products",index=False)
		i.fac_table.to_excel(writer, "Facilities",index=False)
		#sheet = wb['Data1']

		writer.save()
		os.system("open -i 'Microsoft Excel' 'LMIS Analysis Report.xlsm'")
		logging.info('Report created')

	def openSettings(self):
		window2=Tk()
		window2.title("Inventory Analysis Settings")

	def createConfig(self):
		config = configparser.ConfigParser()
		config.add_section('column headers')
		config.set('column headers','data_type', self.data_type.get())
		config.set('column headers','product',self.product.get())
		config.set('column headers','product1',self.product1.get())
		config.set('column headers','product2',self.product2.get())
		config.set('column headers','facility',self.facility.get())
		config.set('column headers','facility1',self.facility1.get())
		config.set('column headers','facility2',self.facility2.get())
		config.set('column headers','date_field',self.date_field.get())
		config.set('column headers','issue_col',self.issue_col.get())
		config.set('column headers','soh_col',self.soh_col.get())

		config.add_section('other')
		config.set('other','rates_window',str(self.rates_window.get()))
		config.set('other','issue_0',str(self.issue_0.get()))
		config.set('other','soh_0',str(self.soh_0.get()))

		basepath = path.dirname(__file__)

		with open(path.join(basepath,'configfile.ini'),'w') as configfile:
			config.write(configfile)


		


IAGui(window,is_config=True)
window.mainloop()

